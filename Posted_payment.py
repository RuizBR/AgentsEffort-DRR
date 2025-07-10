import streamlit as st
import pandas as pd
import mysql.connector
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import date

st.title("🧾 Agent Posted Payments Report")

# --- Date Range Picker ---
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Start Date", value=date(2025, 4, 10))
with col2:
    end_date = st.date_input("End Date", value=date(2025, 4, 14))

# MySQL connection
def create_connection():
    return mysql.connector.connect(
        host="172.16.128.79",
        user="usr4mis",
        password="usr4MIS#@!",
        port=3307,
        database="volare"
    )

# --- Validate Dates ---
if start_date > end_date:
    st.error("🚫 Start date must be before end date.")
else:
    # SQL query
    query = f"""
    SELECT DISTINCT
        CONCAT('Cycle ', RIGHT(debtor.cycle, 2)) AS `CYCLE`,
        debtor.card_no AS `CH CODE`,
        debtor.account AS `ACCOUNT NUMBER`,
        followup.remark AS `REMARKS`,
        followup.remark_by AS `AGENT CODE`,
        followup.status_code AS `STATUS CODE`,
        debtor.ptp_amount AS `PTP AMOUNT`,
        debtor.ptp_date AS `PTP DATE`,
        debtor.balance AS `OB`,
        followup.datetime AS `DISPO DATE`,
        debtor.placement AS `FINONE ID`,
        debtor.is_locked AS `IS LOCKED`,
        debtor.is_aborted AS `IS ABORTED`
    FROM debtor
    LEFT JOIN debtor_followup ON debtor_followup.debtor_id = debtor.id
    LEFT JOIN followup ON followup.id = debtor_followup.followup_id
    LEFT JOIN `user` ON `user`.id = followup.remark_by_id
    WHERE 1=1
        AND debtor.client_name LIKE 'BPI CARDS XDAYS%'
        AND followup.status_code IN ('PAYMENT - CURED')
        AND DATE_FORMAT(followup.`date`, '%Y-%m-%d') BETWEEN '{start_date}' AND '{end_date}'
        AND followup.remark_by NOT IN ('BLRUIZ', 'KPILUSTRISIMO', 'MMMEJIA', 'SAHERNANDEZ', 'FGPANGANIBAN')
        AND followup.remark NOT LIKE '%MSPM%'
    ORDER BY followup.datetime DESC
    """

    try:
        conn = create_connection()
        df = pd.read_sql(query, conn)

        if df.empty:
            st.warning("No data found for the selected date range.")
        else:
            st.success(f"✅ Loaded {len(df)} records.")

            # Format ACCOUNT NUMBER with leading zeros
            if "ACCOUNT NUMBER" in df.columns:
                df["ACCOUNT NUMBER"] = df["ACCOUNT NUMBER"].astype(str).str.zfill(10)

            # Display DataFrame
            st.dataframe(df, use_container_width=True)

            # --- CSV Download ---
            csv = df.to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ Download CSV", csv, "agent_posted_payments.csv", "text/csv")

            # --- Excel Download with autofit columns ---
            export_df = df.copy()
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Agent Payments"

            for r in dataframe_to_rows(export_df, index=False, header=True):
                ws.append(r)

            for col in ws.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            wb.save(output)
            st.download_button(
                label="⬇️ Download Excel",
                data=output.getvalue(),
                file_name="agent_posted_payments.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error loading data: {e}")
