import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# Define header style: black fill, white text, bold, centered
header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_align = Alignment(horizontal="center", vertical="center")

# st.set_page_config(page_title="📄 Daily Efforts per Agent", layout="centered")
st.title("📄 Daily Efforts per Agent")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # Standardize column names
    df.columns = df.columns.str.strip().str.lower().str.replace('\xa0', ' ').str.replace('\n', ' ')

    # Identify required columns
    remark_col = next((col for col in df.columns if "remark by" in col), None)
    status_col = next((col for col in df.columns if col == "status" and "call" not in col), None)
    balance_col = next((col for col in df.columns if col == "balance"), None)

    # Set standard column order for export
    standard_columns = [
        "cycle", "client", "account no.", "card no.", "debtor",
        "call status", "status", "remark", "remark by", "ptp amount",
        "ptp date", "dialed number", "balance", "min payment"
    ]

    if not remark_col or not status_col or not balance_col:
        st.error(f"❌ Required columns missing: 'Remark By', 'Status', or 'Balance'. Found: {list(df.columns)}")
    else:
        remark_by_list = sorted(df[remark_col].dropna().unique())
        selected_remark_by = st.selectbox("👤 Select a 'Remark By'", remark_by_list)

        if selected_remark_by:
            filtered_df = df[df[remark_col] == selected_remark_by].copy()

            # ✅ Clean and convert balance to float
            filtered_df[balance_col] = (
                filtered_df[balance_col]
                .astype(str)
                .str.replace(r"[^\d\.-]", "", regex=True)
                .replace("", "0")
                .astype(float)
            )

            st.markdown(f"🔎 **Filtered rows for `{selected_remark_by}`: {len(filtered_df)}**")
            st.dataframe(filtered_df.head())

            # Define filter conditions for sheets
            sheets = {
                "Bank Escalation": filtered_df[filtered_df[status_col].str.contains("BANK ESCALATION", na=False, case=False)],
                "PTP": filtered_df[filtered_df[status_col].str.contains("PTP", na=False, case=False) & ~filtered_df[status_col].str.contains("PTP FF UP", na=False, case=False)],
                "Payment - Cured": filtered_df[filtered_df[status_col].str.contains("PAYMENT - CURED", na=False, case=False)],
                "Negative": filtered_df[filtered_df[status_col].str.contains("NEGATIVE", na=False, case=False)],
                "RPC": filtered_df[filtered_df[status_col].str.contains("RPC", na=False, case=False)],
                "TPC": filtered_df[filtered_df[status_col].str.contains("TPC", na=False, case=False)],
            }

            output = BytesIO()
            all_pivot_data = []

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for sheet_name, sheet_df in sheets.items():
                    st.markdown(f"📄 `{sheet_name}`: {len(sheet_df)} rows")
                    if not sheet_df.empty:
                        # ✅ Clean balance in each sub-sheet
                        sheet_df[balance_col] = (
                            sheet_df[balance_col]
                            .astype(str)
                            .str.replace(r"[^\d\.-]", "", regex=True)
                            .replace("", "0")
                            .astype(float)
                        )

                        # Reorder columns and write to sheet
                        output_df = sheet_df.reindex(columns=standard_columns, fill_value="")
                        output_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        ws = writer.sheets[sheet_name[:31]]

                        # Style header
                        for col_idx in range(1, len(output_df.columns) + 1):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_align

                        # Auto-fit columns
                        for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
                            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

                        # Create pivot
                        pivot = sheet_df.groupby(["cycle", status_col], dropna=False).agg(
                            count=(status_col, "size"),
                            total_balance=(balance_col, "sum")
                        ).reset_index()

                        pivot.insert(0, "category", sheet_name)
                        pivot = pivot[["category", "cycle", status_col, "count", "total_balance"]]
                        all_pivot_data.append(pivot)

                # ➕ Write all pivots into a single sheet
                summary_ws = writer.book.create_sheet("Summary")
                start_row = 1

                for pivot_df in all_pivot_data:
                    category = pivot_df["category"].iloc[0]

                    # Title row
                    summary_ws.cell(row=start_row, column=1, value=f"{category} Summary")
                    start_row += 1

                    # Styled header row
                    for col_idx, col_name in enumerate(pivot_df.columns, 1):
                        cell = summary_ws.cell(row=start_row, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_align

                    # Data rows
                    for row_idx, row in pivot_df.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            summary_ws.cell(row=start_row + 1 + row_idx, column=col_idx, value=value)

                    # Auto-fit summary columns
                    for col_idx in range(1, len(pivot_df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        max_len = max(
                            len(str(summary_ws.cell(row=r, column=col_idx).value)) if summary_ws.cell(row=r, column=col_idx).value else 0
                            for r in range(start_row, start_row + len(pivot_df) + 1)
                        )
                        summary_ws.column_dimensions[col_letter].width = max_len + 2

                    start_row += len(pivot_df) + 4

            st.success("✅ Excel file with styled headers and summary is ready!")
            st.download_button(
                label="📥 Download Filtered Excel File",
                data=output.getvalue(),
                file_name=f"{selected_remark_by}_Agents-Efforts-Daily.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
